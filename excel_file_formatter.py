# Excel File Formatter
from pathlib import Path 
import openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
print('Excel File Formatter\n===============')
while True:
    file_path = input('Please enter the file path: ').strip()
    if Path(file_path).is_file():
        break
    print('The file does not exist. please try again: ')
folder_path = Path(file_path).parent
file_name = Path(file_path).name
try :
    wb = openpyxl.load_workbook(file_path)
except Exception:
    print("Invalid Excel workbook.")
    sys.exit()
sheet = wb.active
total_row = sheet.max_row +1
sheet[f'A{total_row}'] = 'Total'
summary_headers = ('revenue', 'price', 'cost', 'sold', 'total', 'sales', 'profit')
money_headers = ('sales','price','cost','revenue','profit')
# number formatting & total row
for col in range(1, sheet.max_column + 1):
    header = str(sheet.cell(row= 1, column = col).value).lower()
    if any(word in header for word in money_headers):
        for row in range(2, total_row +1):
            sheet.cell(row =row, column = col).number_format = '$#,##0.00'
    if any(word in header for word in summary_headers):
        sheet.cell(row = total_row, column = col).value = f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row - 1})'
# font, color, text alignment formatting
header_font = Font(size = 12, color = '00FFFFFF', bold = True)
text_alignment = Alignment(horizontal ='center', vertical = 'center')
header_fill = PatternFill(start_color = '1F4E79', fill_type= 'solid')
right_border = Side(border_style= 'thick', color ='7F7F7F')
right_col_border = Border(right= right_border)
# cell content alignment
for col in range(1, sheet.max_column +1):
    for row in range(1, total_row + 1):
        cell = sheet.cell(row = row, column = col)
        cell.alignment = text_alignment
# adjust column width
for each_col in range(1, sheet.max_column +1):
    letter = get_column_letter(each_col)
    sheet.column_dimensions[letter].width = 20
# style header row and total
for i in range(1, sheet.max_column +1):
    sheet.cell(row = 1, column = i).font = header_font
    sheet.cell(row = total_row , column = i).font =Font(bold =True)
    sheet.cell(row = 1, column = i).fill = header_fill
#apply border to first column
for row in range(2, total_row +1):
    sheet.cell(row = row, column = 1).border = right_col_border
sheet[f'A{total_row}'].font = Font(size = 16, bold = True)
sheet.freeze_panes = 'A2'
sheet.auto_filter.ref = sheet.dimensions
wb.save(folder_path / f'updated_{file_name}')
print('\nWorkbook formatted successfully!')
print(f'\nSaved as: updated_{file_name}')
